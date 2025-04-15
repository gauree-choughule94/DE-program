import os
import pandas as pd
from typing import Dict, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pandas import DataFrame
from logger import log_success, log_warning, log_error


CSV_DIR = "csv_files"
EXCEL_DIR = "excel_files"

os.makedirs(CSV_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)


def save_data_in_folders(dataframes: Dict[str, pd.DataFrame]) -> None:
    """
    Saves cleaned data as CSV and Excel files.
    """
    for key, df in dataframes.items():
        if not df.empty:
            try:
                csv_path = os.path.join(CSV_DIR, f"{key}.csv")
                excel_path = os.path.join(EXCEL_DIR, f"{key}.xlsx")

                df.to_csv(csv_path, index=False, encoding="utf-8")
                df.to_excel(excel_path, index=False)

                log_success(f"Saved {key} data to {csv_path} and {excel_path}.")
            except Exception as e:
                log_error(f"Error saving data for '{key}': {e}")
        else:
            log_warning(f"No data found for '{key}'. Skipping save.")


TABLE_MAPPING = {
    "users.csv": "users",
    "payments.csv": "payments",
    "products.csv": "products",
    "details.csv": "details",
}


def insert_csv_data(engine) -> None:
    """
    Reads CSV files from the CSV_DIR and inserts data into the corresponding database tables.
    """
    for file_name, table_name in TABLE_MAPPING.items():
        file_path = os.path.join(CSV_DIR, file_name)

        if os.path.exists(file_path):
            try:
                df = pd.read_csv(file_path)

                if not df.empty:
                    df.to_sql(table_name, con=engine, if_exists="replace", index=False)
                    log_success(f"Inserted {len(df)} records into '{table_name}' table.")
                else:
                    log_warning(f"Skipping '{file_name}' as it is empty.")
            except Exception as e:
                log_error(f"Error inserting data from '{file_name}' into '{table_name}': {e}")
        else:
            log_warning(f"File '{file_name}' not found. Skipping.")


REPORT_FILE = "data_report.xlsx"


def generate_report(dfs: Dict[str, DataFrame], insights: Tuple[Any, Any, float, Any]) -> None:
    """
    Generates an Excel report summarizing raw data and key insights.
    """
    try:
        expensive_users, users_from_cluj, total_price, jpmorgan_users = insights

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        ws["A1"] = "Data Report"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:D1")

        ws["A3"] = "Profit Analysis"
        ws["A3"].font = Font(size=12, bold=True)
        ws["A3"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A3:D3")

        ws["A5"] = "Total Purchase Amount"
        ws["A5"].font = Font(size=12, bold=True)
        ws["B5"] = f"${total_price:.2f}"

        ws["A7"] = "Users who bought products over $25"
        ws["A7"].font = Font(size=12, bold=True)
        ws.merge_cells("A7:D7")

        ws["A8"], ws["B8"] = "User ID", "User Name"
        ws["A8"].font = ws["B8"].font = Font(bold=True)

        row = 9
        for user_id, name in expensive_users:
            ws[f"A{row}"] = user_id
            ws[f"B{row}"] = name
            row += 1

        row += 2
        ws[f"A{row}"] = "Users from Cluj-Napoca"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        ws[f"A{row}"], ws[f"B{row}"] = "User ID", "User Name"
        ws[f"A{row}"].font = ws[f"B{row}"].font = Font(bold=True)
        row += 1
        for user_id, name in users_from_cluj:
            ws[f"A{row}"] = user_id
            ws[f"B{row}"] = name
            row += 1

        row += 2
        ws[f"A{row}"] = "Users who paid via JPMORGAN CHASE"
        ws[f"A{row}"].font = Font(size=12, bold=True)
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        ws[f"A{row}"], ws[f"B{row}"] = "User ID", "User Name"
        ws[f"A{row}"].font = ws[f"B{row}"].font = Font(bold=True)
        row += 1
        for user_id, name in jpmorgan_users:
            ws[f"A{row}"] = user_id
            ws[f"B{row}"] = name
            row += 1

        ws[f"A{row+1}"] = "Total JPMorgan Users"
        ws[f"B{row+1}"] = len(jpmorgan_users)

        wb.save(REPORT_FILE)
        log_success(f"Report saved as {REPORT_FILE}")
        print(f"Report saved as {REPORT_FILE}")

    except Exception as e:
        log_error(f"Error generating report: {e}")


def save_transformed_outputs_to_excel(
    filtered_df: pd.DataFrame,
    appended_df: pd.DataFrame,
    deleted_df: pd.DataFrame,
    union_df: pd.DataFrame,
    aggregated_df: pd.DataFrame,
    file_name: str = "transformed_report.xlsx"
):
    try:
        with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
            filtered_df.to_excel(writer, sheet_name="Filtered Data", index=False)
            appended_df.to_excel(writer, sheet_name="Appended Data", index=False)
            deleted_df.to_excel(writer, sheet_name="Deleted Data", index=False)
            union_df.to_excel(writer, sheet_name="Union Data", index=False)
            aggregated_df.to_excel(writer, sheet_name="Aggregated Data", index=False)

        log_success(f"Transformed data saved to {file_name}")

    except Exception as e:
        log_error(f"Error saving transformed data: {e}")
